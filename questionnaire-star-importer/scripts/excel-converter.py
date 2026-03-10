#!/usr/bin/env python3
"""
问卷星Excel导入格式转换脚本
用法：python3 excel-converter.py <输入文件> <输出文件.xlsx>
"""
import re
import sys
from pathlib import Path
try:
    from openpyxl import Workbook
except ImportError:
    print("⚠️  需要安装openpyxl库，正在自动安装...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import Workbook

def parse_question(content):
    """解析单道题目"""
    lines = content.strip().split("\n")
    if not lines:
        return None
    
    # 第一行是题目
    first_line = lines[0].strip()
    # 去除题号
    q_match = re.match(r"(\d+)[.、\s]+(.+)", first_line)
    if q_match:
        q_no = int(q_match.group(1))
        q_content = q_match.group(2)
    else:
        q_no = 0
        q_content = first_line
    
    # 识别题型
    q_type = "单选题" # 默认单选
    type_match = re.search(r"【(多选题|量表题|填空题|多行填空题|排序题|评分题)(.*)】", q_content)
    if type_match:
        q_type = type_match.group(1)
        q_content = q_content.replace(type_match.group(0), "").strip()
    else:
        # 检测是否是多选：选项超过4个/有"可多选"字样
        if "可多选" in q_content or "多选" in q_content:
            q_type = "多选题"
    
    # 处理选项
    options = []
    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue
        opt_match = re.match(r"([A-Za-z])[.、\s]+(.+)", line)
        if opt_match:
            opt_text = opt_match.group(2).strip()
            options.append(opt_text)
        elif re.match(r"\d+[.、\s]+", line): # 数字开头的选项
            opt_text = re.sub(r"\d+[.、\s]+", "", line).strip()
            options.append(opt_text)
    
    return {
        "no": q_no,
        "content": q_content,
        "type": q_type,
        "options": options
    }

def convert_to_excel(input_path, output_path, title="问卷标题"):
    """转换为问卷星Excel导入格式"""
    content = Path(input_path).read_text(encoding="utf-8")
    
    # 分割题目（按数字题号分割）
    question_blocks = re.split(r"\n\d+[.、]", content)
    if not question_blocks[0].strip():
        question_blocks = question_blocks[1:]
    
    # 解析所有题目
    questions = []
    for idx, q_content in enumerate(question_blocks, 1):
        # 补上题号
        q_content = f"{idx}. {q_content.strip()}"
        parsed_q = parse_question(q_content)
        if parsed_q:
            parsed_q["no"] = idx
            questions.append(parsed_q)
    
    # 生成Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "问卷导入"
    
    # 写入表头
    headers = ["题目序号", "题目内容", "题型", "选项1", "选项2", "选项3", "选项4", "选项5", "选项6", "选项7", "选项8", "选项9", "选项10"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入题目
    for row_idx, q in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=q["no"])
        ws.cell(row=row_idx, column=2, value=q["content"])
        ws.cell(row=row_idx, column=3, value=q["type"])
        
        # 写入选项
        for opt_idx, opt in enumerate(q["options"], 4):
            if opt_idx <= 13: # 最多支持10个选项
                ws.cell(row=row_idx, column=opt_idx, value=opt)
    
    # 调整列宽
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15
    for col in range(4, 14):
        ws.column_dimensions[chr(ord('A') + col - 1)].width = 20
    
    # 保存文件
    wb.save(output_path)
    print(f"✅ Excel导入格式已生成：{output_path}")
    print("💡 直接上传到问卷星「Excel快速导入」即可生成完整问卷")

def main():
    if len(sys.argv) < 3:
        print("用法：python3 excel-converter.py <输入文件> <输出文件.xlsx> [问卷标题]")
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    title = sys.argv[3] if len(sys.argv) > 3 else "问卷标题"
    
    if not Path(input_path).exists():
        print(f"❌ 输入文件不存在：{input_path}")
        sys.exit(1)
    
    convert_to_excel(input_path, output_path, title)

if __name__ == "__main__":
    main()
